"""Génération de fichiers Excel pour l'export des contrats avec le style d'origine."""

from __future__ import annotations
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import InvoiceData

# Couleurs d'origine configurées pour openpyxl
HEADER_FILL_COLOR = "1F4E79"   # Bleu acier d'origine
ALT_ROW_FILL_COLOR = "DCE6F1"  # Bleu très clair alterné
WHITE_COLOR = "FFFFFF"
BORDER_GRAY = "D9D9D9"

def contracts_to_excel(invoice: InvoiceData, plant: str = "", dept: str = "", project: str = "") -> bytes:
    """Génère un fichier Excel calqué fidèlement sur l'extraction d'origine, sans bloc de métadonnées."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Factures"  # Onglet calqué sur l'ancien code

    # Nettoyage du numéro de facture pour ne garder que le numéro numérique (Retire ZONE AGROPOLIS, etc.)
    raw_invoice_num = str(invoice.invoice_number or '')
    clean_invoice_num = raw_invoice_num.split(" ")[0].strip() if " " in raw_invoice_num else raw_invoice_num

    # Définition des styles standards
    hdr_fill = PatternFill("solid", start_color=HEADER_FILL_COLOR)
    alt_fill = PatternFill("solid", start_color=ALT_ROW_FILL_COLOR)
    white_fill = PatternFill("solid", start_color=WHITE_COLOR)
    
    font_header = Font(name="Arial", bold=True, color=WHITE_COLOR, size=11)
    font_data = Font(name="Arial", size=10)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_side = Side(style="thin", color=BORDER_GRAY)
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # L'en-tête du tableau commence désormais directement à la Ligne 1 (sans métadonnées)
    start_table_row = 1

    # ── 1. En-têtes de Colonnes de la Grille (Ligne 1) ─────────────────────
    headers_display = [
        "Page PDF", "N° Client", "N° Facture", "N° d'Appel (SIM)", 
        "Date Facture", "Début Période", "Fin Période", "Total Contrat (MAD)"
    ]
    
    # Largeurs de colonnes de l'ancienne configuration
    col_widths = [10, 24, 22, 18, 14, 14, 14, 20]

    for ci, (hdr_text, width) in enumerate(zip(headers_display, col_widths), start=1):
        c = ws.cell(row=start_table_row, column=ci, value=hdr_text)
        c.font = font_header
        c.fill = hdr_fill
        c.alignment = align_center
        c.border = cell_border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[start_table_row].height = 26

    # ── 2. Remplissage des Lignes de Données Facture (Lignes 2+) ────────────
    current_row = start_table_row + 1
    
    for ri, contract in enumerate(invoice.contracts, start=current_row):
        # Alternance des couleurs (Zebra rows style)
        row_fill = alt_fill if ri % 2 == 0 else white_fill
        
        # Extraction ordonnée en utilisant le numéro de facture nettoyé
        row_values = [
            contract.page_number,
            invoice.client_number or "—",
            clean_invoice_num or "—",
            contract.phone_number or "—",
            invoice.invoice_date or "—",
            contract.period_start or invoice.period_start or "—",
            contract.period_end or invoice.period_end or "—",
            contract.total_contrat
        ]
        
        for ci, val in enumerate(row_values, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = font_data
            cell.fill = row_fill
            cell.border = cell_border
            
            # Alignements et formats par type de cellule
            if ci in [1, 4, 5, 6, 7]:  # Pages, numéros de tel, dates
                cell.alignment = align_center
            elif ci == 8:              # Montants financiers des contrats
                cell.number_format = "#,##0.00"
                cell.alignment = align_right
            else:
                cell.alignment = align_left
                
        ws.row_dimensions[ri].height = 20
        current_row = ri

    # ── 3. Ligne de Clôture Budgétaire / Formule de Totalisation NATIVE ─────
    total_row = current_row + 1
    ws.merge_cells(f"A{total_row}:G{total_row}")
    
    # Label TOTAL
    total_label_cell = ws.cell(row=total_row, column=1, value="TOTAL DES ENVELOPPES IMPUTÉES :")
    total_label_cell.font = font_header
    total_label_cell.fill = PatternFill("solid", start_color="000000")
    total_label_cell.alignment = align_right
    
    # Évite les pertes de bordures noires sur les cellules fusionnées
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).fill = PatternFill("solid", start_color="000000")

    # Injection dynamique de la formule Excel native SUM
    total_value_cell = ws.cell(row=total_row, column=8)
    total_value_cell.value = f"=SUM(H{start_table_row + 1}:H{total_row - 1})"
    total_value_cell.font = font_header
    total_value_cell.fill = PatternFill("solid", start_color="000000")
    total_value_cell.number_format = '#,##0.00" MAD"'
    total_value_cell.alignment = align_right
    total_value_cell.border = cell_border
    
    ws.row_dimensions[total_row].height = 26

    # Sauvegarde finale dans le flux binaire
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def contracts_to_excel_batch(invoices: list[InvoiceData], plant: str = "", dept: str = "", project: str = "") -> list[bytes]:
    """Génère individuellement chaque fichier Excel et retourne une liste de flux binaires (bytes)."""
    excel_files: list[bytes] = []

    for invoice in invoices:
        # Appel direct de la fonction unitaire pour chaque facture de la liste
        excel_data = contracts_to_excel(invoice, plant=plant, dept=dept, project=project)
        excel_files.append(excel_data)

    return excel_files